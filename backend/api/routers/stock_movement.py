from fastapi import APIRouter, HTTPException, Depends
from typing import Dict, List, Optional, Any, Set
from pydantic import BaseModel
import pandas as pd
import numpy as np
import os
from datetime import datetime

router = APIRouter()

class UOMMapping(BaseModel):
    """UOM mapping model"""
    uom_name: str
    units: int

class StockMovementConfig(BaseModel):
    """Configuration for stock movement processing"""
    session_id: str
    uom_column: str  # Column containing UOM data (e.g., "Unit of Measure")
    location_from_column: str  # Column for source location (e.g., "From")
    location_to_column: str  # Column for destination location (e.g., "To")
    quantity_column: str  # Column for quantity (e.g., "Done")
    check_location: str  # Location to check for sign calculation (e.g., "JDFCW/Jeddah Factory")
    uom_mappings: List[UOMMapping]  # List of UOM to units mappings
    date_column: Optional[str] = None  # Optional date column for sorting

class StockMovementResult(BaseModel):
    """Result of stock movement processing"""
    success: bool
    message: str
    processed_data: Optional[List[Dict]] = None
    summary: Optional[Dict] = None
    columns: Optional[List[str]] = None

def get_session_file(session_id: str) -> str:
    """Get the file path for a session"""
    # Check uploads directory
    uploads_dir = "uploads"
    for file in os.listdir(uploads_dir):
        if file.startswith(session_id):
            return os.path.join(uploads_dir, file)
    
    raise HTTPException(status_code=404, detail="Session file not found")


def validate_stock_movement_file(df: pd.DataFrame) -> bool:
    """Validate if the file is a valid stock movement report"""
    # Required columns for stock movement report
    required_columns = {
        'Date', 'Reference', 'Product', 'From', 'To', 'Done', 'Unit of Measure', 'Status'
    }
    
    # Check if all required columns are present
    existing_columns = set(df.columns)
    missing_columns = required_columns - existing_columns
    
    if missing_columns:
        return False
    
    # Additional validation can be added here if needed
    return True

@router.post("/process", response_model=StockMovementResult)
async def process_stock_movement(config: StockMovementConfig):
    """
    Process stock movement data with UOM conversion and running balance
    """
    try:
        # Get the file path
        file_path = get_session_file(config.session_id)
        
        # Read the data
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        # Validate if this is a stock movement file
        if not validate_stock_movement_file(df):
            return StockMovementResult(
                success=False,
                message="The uploaded file does not appear to be a valid stock movement report. Please check the file format."
            )
        
        # Create UOM mapping dictionary
        uom_dict = {mapping.uom_name: mapping.units for mapping in config.uom_mappings}
        
        # Process UOM Qty column
        if config.uom_column in df.columns:
            def convert_uom(value):
                """Convert UOM to quantity"""
                if pd.isna(value):
                    return 1
                value_str = str(value)
                for uom_name, units in uom_dict.items():
                    if uom_name in value_str:
                        return units
                return 1
            
            df['UOM_Qty_Done'] = df[config.uom_column].apply(convert_uom)
        else:
            df['UOM_Qty_Done'] = 1
        
        # Process Sign column based on location (To/From)
        if config.location_from_column in df.columns and config.location_to_column in df.columns and config.quantity_column in df.columns:
            def calculate_sign(row):
                """Calculate sign based on location (inbound/outbound)"""
                from_location = str(row[config.location_from_column])
                to_location = str(row[config.location_to_column])
                quantity = row[config.quantity_column]
                
                # If destination is the check location, it's an inbound movement (+)
                if config.check_location in to_location:
                    return quantity
                # If source is the check location, it's an outbound movement (-)
                elif config.check_location in from_location:
                    return -quantity
                # For other movements, use destination as positive
                else:
                    return quantity
            
            df['Sign'] = df.apply(calculate_sign, axis=1)
        else:
            df['Sign'] = df[config.quantity_column] if config.quantity_column in df.columns else 0
        
        # Apply UOM conversion to Sign column (Cartons x Units x Sign = UOM_Done_Qty)
        df['UOM_Done_Qty'] = df['Sign'] * df['UOM_Qty_Done']
        
        # Sort by date if date column is provided
        if config.date_column and config.date_column in df.columns:
            df[config.date_column] = pd.to_datetime(df[config.date_column], errors='coerce')
            df = df.sort_values(by=config.date_column)
        
        # Calculate running balance
        df['Running_Balance'] = df['UOM_Done_Qty'].cumsum()
        
        # Prepare the result
        result_data = df.to_dict('records')
        
        # Calculate summary statistics
        summary = {
            'total_rows': len(df),
            'total_inbound': float(df[df['Sign'] > 0]['UOM_Done_Qty'].sum()) if 'UOM_Done_Qty' in df.columns else 0,
            'total_outbound': float(df[df['Sign'] < 0]['UOM_Done_Qty'].sum()) if 'UOM_Done_Qty' in df.columns else 0,
            'final_balance': float(df['Running_Balance'].iloc[-1]) if len(df) > 0 else 0,
            'unique_from_locations': df[config.location_from_column].nunique() if config.location_from_column in df.columns else 0,
            'unique_to_locations': df[config.location_to_column].nunique() if config.location_to_column in df.columns else 0,
            'unique_uoms': df[config.uom_column].nunique() if config.uom_column in df.columns else 0
        }
        
        return StockMovementResult(
            success=True,
            message="Stock movement processed successfully",
            processed_data=result_data,
            summary=summary,
            columns=list(df.columns)
        )
        
    except Exception as e:
        return StockMovementResult(
            success=False,
            message=f"Error processing stock movement: {str(e)}"
        )

@router.get("/validate/{session_id}")
async def validate_stock_movement_file_endpoint(session_id: str):
    """
    Validate if the uploaded file is a valid stock movement report
    """
    try:
        file_path = get_session_file(session_id)
        
        # Read the data
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        
        is_valid = validate_stock_movement_file(df)
        
        result = {
            "success": True,
            "is_valid_stock_movement": is_valid,
            "columns": list(df.columns),
            "shape": df.shape
        }
        
        if is_valid:
            result["message"] = "File is a valid stock movement report"
            # Add sample data to help with column identification
            result["sample_data"] = df.head(3).to_dict('records')
        else:
            result["message"] = "File does not appear to be a valid stock movement report"
            required_columns = {'Date', 'Reference', 'Product', 'From', 'To', 'Done', 'Unit of Measure', 'Status'}
            missing_columns = required_columns - set(df.columns)
            result["missing_columns"] = list(missing_columns)
        
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/columns/{session_id}")
async def get_available_columns(session_id: str):
    """
    Get available columns from the uploaded file
    """
    try:
        file_path = get_session_file(session_id)
        
        # Read just the headers
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, nrows=1)
        else:
            df = pd.read_excel(file_path, nrows=1)
        
        return {
            "success": True,
            "columns": list(df.columns),
            "column_types": {col: str(df[col].dtype) for col in df.columns}
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/export/{session_id}")
async def export_processed_data(session_id: str, config: StockMovementConfig):
    """
    Export processed stock movement data with new columns
    """
    try:
        # Process the data first
        result = await process_stock_movement(config)
        
        if not result.success:
            raise HTTPException(status_code=400, detail=result.message)
        
        # Convert to DataFrame
        df = pd.DataFrame(result.processed_data)
        
        # Generate export filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_filename = f"stock_movement_{session_id}_{timestamp}.xlsx"
        export_path = f"uploads/{export_filename}"
        
        # Save to Excel with formatting
        with pd.ExcelWriter(export_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Stock Movement', index=False)
            
            # Get the worksheet
            worksheet = writer.sheets['Stock Movement']
            
            # Auto-adjust column widths
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            for idx, column in enumerate(df.columns, 1):
                column_length = max(df[column].astype(str).map(len).max(), len(column))
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = min(column_length + 2, 50)
                
                # Style the header
                header_cell = worksheet[f'{col_letter}1']
                header_cell.font = Font(bold=True)
                header_cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
                header_cell.alignment = Alignment(horizontal='center')
                
                # Highlight new calculated columns
                if column in ['UOM_Qty_Done', 'Sign', 'UOM_Done_Qty', 'Running_Balance']:
                    header_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            
            # Apply number formatting to numeric columns
            for idx, column in enumerate(df.columns, 1):
                col_letter = get_column_letter(idx)
                if column in ['Done', 'UOM_Qty_Done', 'Sign', 'UOM_Done_Qty', 'Running_Balance']:
                    for row in range(2, len(df) + 2):
                        worksheet[f'{col_letter}{row}'].number_format = '#,##0.00'
        
        return {
            "success": True,
            "message": "Data exported successfully",
            "filename": export_filename,
            "path": export_path
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

#!/usr/bin/env python3
"""
Test script for Excel export functionality
"""

import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def test_excel_export():
    """Test the Excel export logic"""
    
    # Create sample processed data
    data = [
        {
            'Date': '2024-12-19 17:13:23',
            'Reference': 'JDFCW/RET/02960',
            'Product': 'Arjoon Mamool Strawberry Flv 150 gms',
            'From': 'Partners/Customers',
            'To': 'JDFCW/Jeddah Factory',
            'Done': 48.0,
            'Unit of Measure': 'Units',
            'UOM_Qty_Done': 1,
            'Sign': 48.0,
            'UOM_Done_Qty': 48.0,
            'Running_Balance': 48.0
        },
        {
            'Date': '2024-12-22 16:40:00',
            'Reference': 'JDFCW/OUT/16346',
            'Product': 'Arjoon Mamool Strawberry Flv 150 gms',
            'From': 'JDFCW/Jeddah Factory',
            'To': 'Partners/Customers',
            'Done': 0.9917,
            'Unit of Measure': 'Cartons-24',
            'UOM_Qty_Done': 24,
            'Sign': -0.9917,
            'UOM_Done_Qty': -23.8008,
            'Running_Balance': 24.1992
        }
    ]
    
    df = pd.DataFrame(data)
    
    print("Test Data:")
    print(df)
    print()
    
    # Create Excel file in memory
    excel_buffer = io.BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Stock Movement', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Stock Movement']
        
        # Auto-adjust column widths and apply formatting
        for idx, column in enumerate(df.columns, 1):
            column_length = max(df[column].astype(str).map(len).max(), len(column))
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(column_length + 2, 50)
            
            # Style the header
            header_cell = worksheet[f'{col_letter}1']
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            header_cell.alignment = Alignment(horizontal='center')
            
            # Highlight new calculated columns
            if column in ['UOM_Qty_Done', 'Sign', 'UOM_Done_Qty', 'Running_Balance']:
                header_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        # Apply number formatting to numeric columns
        for idx, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(idx)
            if column in ['Done', 'UOM_Qty_Done', 'Sign', 'UOM_Done_Qty', 'Running_Balance']:
                for row in range(2, len(df) + 2):
                    worksheet[f'{col_letter}{row}'].number_format = '#,##0.00'
    
    excel_buffer.seek(0)
    
    # Save the test file
    with open('test_stock_movement_export.xlsx', 'wb') as f:
        f.write(excel_buffer.read())
    
    print("Excel export test completed successfully!")
    print("File saved as: test_stock_movement_export.xlsx")
    print(f"File size: {len(excel_buffer.getvalue())} bytes")

if __name__ == "__main__":
    test_excel_export()
